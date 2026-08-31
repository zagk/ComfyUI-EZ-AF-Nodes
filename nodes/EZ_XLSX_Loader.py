from server import PromptServer  # type: ignore // ComfyUI Core
import os
import io
import base64
import random
from aiohttp import web

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

root_dir = os.path.dirname(os.path.abspath(__file__))
xlsx_path = os.path.abspath(os.path.join(root_dir, "../data/xlsx"))

VALID_EXTENSIONS = (".xlsx", ".xlsm")

# Keyword used to auto-detect the "text" column when not explicitly set.
TEXT_COLUMN_KEYWORDS = ["prompt"]


def _blank_image_tensor():
    """Returns a small black placeholder IMAGE tensor so downstream nodes never receive None."""
    import torch
    return torch.zeros((1, 64, 64, 3), dtype=torch.float32)


def _load_image_tensor_from_bytes(data):
    import torch
    import numpy as np
    with Image.open(io.BytesIO(data)) as img:
        img = img.convert("RGB")
        arr = np.array(img).astype(np.float32) / 255.0
    return torch.from_numpy(arr)[None, ...]


def resolve_column_index(headers, override="", keywords=None):
    """Finds a column index by exact (case-insensitive) name match if `override` is given,
    otherwise falls back to a case-insensitive keyword search across the headers."""
    if override:
        override_l = override.strip().lower()
        for i, h in enumerate(headers):
            if str(h).strip().lower() == override_l:
                return i
        return None
    if keywords:
        for i, h in enumerate(headers):
            hl = str(h).strip().lower()
            for kw in keywords:
                if hl == kw or kw in hl:
                    return i
    return None


def _resolve_sheet(wb, sheet_name=""):
    sheets = wb.sheetnames
    if sheet_name and sheet_name in sheets:
        return wb[sheet_name]
    return wb[sheets[0]] if sheets else wb.active


class EZ_XLSX_Loader:
    @classmethod
    def INPUT_TYPES(cls):
        global xlsx_path
        try:
            xlsx_files = []
            for root, dirs, files in os.walk(xlsx_path):
                for f in files:
                    if f.lower().endswith(VALID_EXTENSIONS):
                        full_path = os.path.join(root, f)
                        rel_path = os.path.relpath(full_path, xlsx_path)
                        xlsx_files.append(rel_path)
        except Exception as e:
            xlsx_files = []

        return {
            "required": {
                "xlsx_file": (xlsx_files, {"tooltip": "XLSX file to search for text by row"}),
                "selection_mode": (["single", "multiple", "random"], {"default": "single", "tooltip":
                                                                        "- single: Allows selection of one item at a time.\n"
                                                                        "- multiple: Allows selection of multiple items. Output will be comma-separated.\n"
                                                                        "- random: Allows selection of multiple items. Randomly outputs one of the selected items on each prompt queue."
                                                                        " Will select from all visible (filtered) items if none or single item is selected."
                                                                        " Uses seed if opt_seed is connected. Always re-executes node if opt_seed is not connected"}),
            },
            "optional": {
                "sheet_name": ("STRING", {"default": "", "tooltip": "Name of the sheet to read. Leave empty to use the first (active) sheet."}),
                "text_column": ("STRING", {"default": "", "tooltip": "Column header whose value becomes the main STRING output (e.g. 'Prompt').\nLeave empty to auto-detect a column named 'Prompt'.\nIf no such column is found, all columns are included, labeled, as before."}),
                "opt_seed": ("INT", {"default": 0, "min": 0, "max": 0xffffffffffffffff, "forceInput": True, "tooltip": "Control SEED, used only in 'random' selection mode.\nIf not connected, always re-executes node on prompt"}),
                "filter_text": ("STRING", {"default": "", "tooltip": "Filter items based on a text string"}),
                "selected_row": ("STRING", {"default": ""}),
            }
        }

    RETURN_TYPES = ("STRING", "STRING", "STRING", "IMAGE")
    RETURN_NAMES = ("STRING", "OPT_FILEPATH", "BATCH_SELECTED", "IMAGE")
    OUTPUT_IS_LIST = (False, False, True, False)
    OUTPUT_TOOLTIPS = (
        "Content of selected row(s).\nIf a 'text_column' (e.g. Prompt) is found, only that column's value is returned.",
        "Path to currently selected xlsx file.",
        "List of all selected items.\nWill output all visible (filtered) items if none or single item is selected.",
        "Image embedded in the xlsx cell next to the first selected row.\nA blank placeholder image is returned if that row has no embedded image.",
    )

    FUNCTION = "browse_xlsx"

    CATEGORY = "EZ NODES"
    DESCRIPTION = "Loads rows from selected xlsx file based on UI selection.\nFirst column is always used to provide labels for UI.\nAuto-detects a 'Prompt' column for text output. Images embedded directly in the xlsx (inserted pictures) are read and returned via the IMAGE output."

    def browse_xlsx(self, xlsx_file, selection_mode="single", sheet_name="", text_column="",
                     selected_row="", filter_text="", opt_seed=0):
        global xlsx_path
        xlsx_file = os.path.join(xlsx_path, xlsx_file)
        xlsx_file = os.path.abspath(xlsx_file)

        if not os.path.isfile(xlsx_file):
            return ("No XLSX file found", xlsx_file, [], _blank_image_tensor())

        if not OPENPYXL_AVAILABLE:
            return ("openpyxl is not installed. Run: pip install openpyxl", xlsx_file, [], _blank_image_tensor())

        data = get_rows_from_xlsx(xlsx_file, sheet_name=sheet_name, filter_text=filter_text)
        headers = data["headers"]
        rows = data["rows"]
        orig_row_indices = data["orig_row_indices"]

        if not headers and not rows:
            return ("No data rows found in file", xlsx_file, [], _blank_image_tensor())

        if not rows:
            return ("No matching rows found", xlsx_file, [], _blank_image_tensor())

        text_col_idx = resolve_column_index(headers, text_column, TEXT_COLUMN_KEYWORDS)

        # Handle selection
        selected_indices = []
        if selection_mode == "random":
            # Use seed for deterministic random selection only if opt_seed is provided and not 0
            if opt_seed is not None and opt_seed != 0:
                random.seed(opt_seed)
            selected_indices = [random.randint(0, len(rows) - 1)]
        elif selection_mode == "multiple":
            if selected_row:
                try:
                    selected_indices = [int(idx) for idx in selected_row.split(",") if idx.strip().isdigit() and int(idx) < len(rows)]
                except Exception:
                    selected_indices = []
        else:  # single
            if not selected_row or not selected_row.isdigit() or int(selected_row) >= len(rows):
                selected_indices = [0]
            else:
                selected_indices = [int(selected_row)]

        def format_row(idx):
            row = rows[idx]
            if text_col_idx is not None and text_col_idx < len(row):
                return str(row[text_col_idx]).strip()
            # Fallback: no dedicated text column found, include all columns labeled.
            out = ""
            for h, v in zip(headers, row):
                out += f"{h}:\n{v}\n\n"
            return out.strip()

        # Format output for main output
        outputs = [format_row(idx) for idx in selected_indices]
        output_str = "\n---\n".join(outputs)

        # Format output for ALL_SELECTED_ROWS (list)
        if len(selected_indices) <= 1:
            all_indices = list(range(len(rows)))
        else:
            all_indices = selected_indices
        all_outputs = [format_row(idx) for idx in all_indices]

        # Load the image embedded in the xlsx for the first selected row, if any.
        image_tensor = _blank_image_tensor()
        if PIL_AVAILABLE and selected_indices:
            first_idx = selected_indices[0]
            if first_idx < len(orig_row_indices):
                orig_idx = orig_row_indices[first_idx]
                row_images = get_row_images(xlsx_file, sheet_name=sheet_name)
                img_bytes = row_images.get(orig_idx)
                if img_bytes:
                    try:
                        image_tensor = _load_image_tensor_from_bytes(img_bytes)
                    except Exception as e:
                        print(f"Error decoding embedded image for row {first_idx}: {e}")

        return (output_str, xlsx_file, all_outputs, image_tensor)

    @classmethod
    def IS_CHANGED(cls, xlsx_file, selection_mode, sheet_name="", text_column="",
                    selected_row="", filter_text="", opt_seed=0):
        if selection_mode == "random":
            # For random mode, include seed in the hash only if opt_seed is provided and not 0
            if opt_seed is not None and opt_seed != 0:
                return str(opt_seed) + str(xlsx_file) + str(selection_mode) + str(filter_text) + str(sheet_name)
            else:
                return float('nan')  # Fall back to normal random behavior
        return selected_row + str(xlsx_file) + str(selection_mode) + str(sheet_name) + str(text_column)

    @classmethod
    def VALIDATE_INPUTS(cls, xlsx_file, selection_mode="single", sheet_name="", text_column="",
                         selected_row="", filter_text="", opt_seed=0):
        global xlsx_path
        xlsx_file = os.path.join(xlsx_path, xlsx_file)
        xlsx_file = os.path.abspath(xlsx_file)
        if not os.path.isfile(xlsx_file):
            return "XLSX file does not exist"
        if not OPENPYXL_AVAILABLE:
            return "openpyxl is not installed. Run: pip install openpyxl"
        return True


def get_directory_structure(path):
    structure = {"name": os.path.basename(path), "children": [], "path": path, "expanded": False}
    try:
        with os.scandir(path) as entries:
            for entry in entries:
                if entry.is_dir():
                    structure["children"].append(get_directory_structure(entry.path))
    except PermissionError:
        pass
    return structure


def _cell_to_str(value):
    if value is None:
        return ""
    return str(value)


def get_sheet_names(file_path):
    if not OPENPYXL_AVAILABLE:
        return []
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            return wb.sheetnames
        finally:
            wb.close()
    except Exception as e:
        print(f"Error reading sheet names from {file_path}: {e}")
        return []


def get_rows_from_xlsx(file_path, sheet_name="", filter_text=""):
    """Reads header/data rows using the fast read-only mode.
    Also returns `orig_row_indices`: the 0-indexed worksheet row number for each
    entry in `rows`, so embedded images (which are anchored to worksheet rows)
    can be matched back to a row even after empty rows are filtered out."""
    if not OPENPYXL_AVAILABLE:
        return {"headers": [], "rows": [], "orig_row_indices": [], "sheets": []}
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheets = wb.sheetnames
            ws = _resolve_sheet(wb, sheet_name)

            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append([_cell_to_str(cell) for cell in row])

            if not all_rows or len(all_rows) < 2:
                return {"headers": [], "rows": [], "orig_row_indices": [], "sheets": sheets}

            headers = all_rows[0]
            # data rows start at worksheet 0-indexed row 1 (row 0 is the header)
            data_rows = list(enumerate(all_rows[1:], start=1))

            # Filter out empty rows (rows where all cells are empty or whitespace)
            data_rows = [(oi, row) for oi, row in data_rows if any(cell.strip() for cell in row)]

            if filter_text:
                data_rows = [(oi, row) for oi, row in data_rows if any(filter_text.lower() in str(cell).lower() for cell in row)]

            rows = [row for oi, row in data_rows]
            orig_row_indices = [oi for oi, row in data_rows]

            return {"headers": headers, "rows": rows, "orig_row_indices": orig_row_indices, "sheets": sheets}
        finally:
            wb.close()
    except Exception as e:
        print(f"Error reading XLSX file {file_path}: {e}")
        return {"headers": [], "rows": [], "orig_row_indices": [], "sheets": []}


def get_row_images(file_path, sheet_name=""):
    """Extracts images embedded directly in the worksheet (e.g. inserted pictures),
    returning {worksheet_row_0indexed: raw_image_bytes}.
    Embedded-image access requires a normal (non read-only) workbook load."""
    if not OPENPYXL_AVAILABLE or not PIL_AVAILABLE:
        return {}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        try:
            ws = _resolve_sheet(wb, sheet_name)
            mapping = {}
            for img in getattr(ws, "_images", []):
                try:
                    anchor_row = img.anchor._from.row  # 0-indexed worksheet row
                except AttributeError:
                    continue
                if anchor_row in mapping:
                    continue  # keep the first image found for a given row
                try:
                    mapping[anchor_row] = img._data()
                except Exception:
                    continue
            return mapping
        finally:
            wb.close()
    except Exception as e:
        print(f"Error extracting embedded images from {file_path}: {e}")
        return {}


@PromptServer.instance.routes.post("/ez_xlsx_browser/get_directory_structure")
async def api_get_directory_structure_xlsx(request):
    try:
        data = await request.json()
        path = data.get("path", "./")
        filter_text = data.get("filter", "")
        sheet_name = data.get("sheet", "")

        if not os.path.isabs(path):
            path = os.path.abspath(path)

        if not os.path.exists(path):
            return web.json_response({"error": "Path does not exist"}, status=400)

        # If path is a file, get its directory
        if os.path.isfile(path):
            directory = os.path.dirname(path)
            structure = get_directory_structure(directory)
            xlsx_data = get_rows_from_xlsx(path, sheet_name=sheet_name, filter_text=filter_text)
        else:
            structure = get_directory_structure(path)
            xlsx_data = {"headers": [], "rows": [], "orig_row_indices": [], "sheets": []}

        response_data = {
            "structure": structure,
            "headers": xlsx_data["headers"],
            "rows": xlsx_data["rows"],
            "orig_row_indices": xlsx_data.get("orig_row_indices", []),
            "sheets": xlsx_data.get("sheets", []),
        }
        return web.json_response(response_data)
    except Exception as e:
        return web.json_response({"error": str(e)}, status=500)


@PromptServer.instance.routes.post("/ez_xlsx_browser/get_file_info")
async def get_file_info_xlsx(request):
    try:
        data = await request.json()
        rel_path = data.get("relative_path", "")
        full_path = os.path.normpath(os.path.join(xlsx_path, rel_path))
        if not full_path.startswith(xlsx_path):
            return web.json_response({"error": "Invalid path"}, status=400)
        if not os.path.exists(full_path):
            return web.json_response({"error": "File not found"}, status=404)
        return web.json_response({
            "full_path": full_path,
            "sheets": get_sheet_names(full_path),
        })
    except Exception as e:
        return web.json_response({"error": str(e)}, status=500)


@PromptServer.instance.routes.post("/ez_xlsx_browser/get_thumbnail")
async def api_get_thumbnail_xlsx(request):
    """Returns a small PNG thumbnail of the image embedded in the worksheet at the
    given row, used by the node's UI to preview the currently selected row's image."""
    try:
        data = await request.json()
        xlsx_full_path = data.get("xlsx_path", "")
        sheet_name = data.get("sheet", "")
        orig_row_index = data.get("orig_row_index", None)

        if not xlsx_full_path or orig_row_index is None:
            return web.json_response({"error": "Missing parameters"}, status=400)

        if not PIL_AVAILABLE:
            return web.json_response({"error": "Pillow is not installed"}, status=500)

        row_images = get_row_images(xlsx_full_path, sheet_name=sheet_name)
        img_bytes = row_images.get(int(orig_row_index))
        if not img_bytes:
            return web.json_response({"error": "No embedded image for this row"}, status=404)

        with Image.open(io.BytesIO(img_bytes)) as img:
            img = img.convert("RGB")
            img.thumbnail((200, 200))
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            buf.seek(0)
            return web.Response(body=buf.read(), content_type="image/png")
    except Exception as e:
        return web.json_response({"error": str(e)}, status=500)


@PromptServer.instance.routes.post("/ez_xlsx_browser/get_thumbnails_batch")
async def api_get_thumbnails_batch_xlsx(request):
    """Returns small base64-encoded thumbnails for every image embedded in the
    worksheet in a single call, keyed by worksheet row (0-indexed) as a string.
    Used to render the node's row list as an image gallery instead of plain text."""
    try:
        data = await request.json()
        xlsx_full_path = data.get("xlsx_path", "")
        sheet_name = data.get("sheet", "")
        size = int(data.get("size", 128))

        if not xlsx_full_path:
            return web.json_response({"error": "Missing xlsx_path"}, status=400)
        if not PIL_AVAILABLE:
            return web.json_response({"error": "Pillow is not installed"}, status=500)

        row_images = get_row_images(xlsx_full_path, sheet_name=sheet_name)
        thumbnails = {}
        for orig_idx, img_bytes in row_images.items():
            try:
                with Image.open(io.BytesIO(img_bytes)) as img:
                    img = img.convert("RGB")
                    img.thumbnail((size, size))
                    buf = io.BytesIO()
                    img.save(buf, format="JPEG", quality=72)
                    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
                    thumbnails[str(orig_idx)] = f"data:image/jpeg;base64,{b64}"
            except Exception as e:
                print(f"Error building thumbnail for row {orig_idx}: {e}")
                continue

        return web.json_response({"thumbnails": thumbnails})
    except Exception as e:
        return web.json_response({"error": str(e)}, status=500)
