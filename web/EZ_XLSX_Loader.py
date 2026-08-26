from server import PromptServer  # type: ignore // ComfyUI Core
import os
import random
from aiohttp import web
import json

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

root_dir = os.path.dirname(os.path.abspath(__file__))
xlsx_path = os.path.abspath(os.path.join(root_dir, "../data/xlsx"))

VALID_EXTENSIONS = (".xlsx", ".xlsm")


class EZ_XLSX_Loader:
    @classmethod
    def INPUT_TYPES(cls):
        global xlsx_path
        try:
            xlsx_files = []
            for root, dirs, files in os.walk(xlsx_path):
                for f in files:
                    if f.lower().endswith(VALID_EXTENSIONS):
                        full_path = os.path.join(root, f)
                        rel_path = os.path.relpath(full_path, xlsx_path)
                        xlsx_files.append(rel_path)
        except Exception as e:
            xlsx_files = []

        return {
            "required": {
                "xlsx_file": (xlsx_files, {"tooltip": "XLSX file to search for text by row"}),
                "selection_mode": (["single", "multiple", "random"], {"default": "single", "tooltip":
                                                                        "- single: Allows selection of one item at a time.\n"
                                                                        "- multiple: Allows selection of multiple items. Output will be comma-separated.\n"
                                                                        "- random: Allows selection of multiple items. Randomly outputs one of the selected items on each prompt queue."
                                                                        " Will select from all visible (filtered) items if none or single item is selected."
                                                                        " Uses seed if opt_seed is connected. Always re-executes node if opt_seed is not connected"}),
            },
            "optional": {
                "sheet_name": ("STRING", {"default": "", "tooltip": "Name of the sheet to read. Leave empty to use the first (active) sheet."}),
                "opt_seed": ("INT", {"default": 0, "min": 0, "max": 0xffffffffffffffff, "forceInput": True, "tooltip": "Control SEED, used only in 'random' selection mode.\nIf not connected, always re-executes node on prompt"}),
                "filter_text": ("STRING", {"default": "", "tooltip": "Filter items based on a text string"}),
                "selected_row": ("STRING", {"default": ""}),
            }
        }

    RETURN_TYPES = ("STRING", "STRING", "STRING")
    RETURN_NAMES = ("STRING", "OPT_FILEPATH", "BATCH_SELECTED")
    OUTPUT_IS_LIST = (False, False, True)
    OUTPUT_TOOLTIPS = ("Content of selected row(s).\nUsually needs to be processed by Prompt Extractor", "Path to currently selected xlsx file.", "List of all selected items.\nWill output all visible (filtered) items if none or single item is selected.")

    FUNCTION = "browse_xlsx"

    CATEGORY = "EZ NODES"
    DESCRIPTION = "Loads rows from selected xlsx file based on UI selection.\nFirst column is always used to provide labels for UI."

    def browse_xlsx(self, xlsx_file, selection_mode="single", sheet_name="", selected_row="", filter_text="", opt_seed=0):
        global xlsx_path
        xlsx_file = os.path.join(xlsx_path, xlsx_file)
        xlsx_file = os.path.abspath(xlsx_file)

        if not os.path.isfile(xlsx_file):
            return ("No XLSX file found", xlsx_file, [])

        if not OPENPYXL_AVAILABLE:
            return ("openpyxl is not installed. Run: pip install openpyxl", xlsx_file, [])

        data = get_rows_from_xlsx(xlsx_file, sheet_name=sheet_name, filter_text=filter_text)
        headers = data["headers"]
        rows = data["rows"]

        if not headers and not rows:
            return ("No data rows found in file", xlsx_file, [])

        if not rows:
            return ("No matching rows found", xlsx_file, [])

        # Handle selection
        selected_indices = []
        if selection_mode == "random":
            # Use seed for deterministic random selection only if opt_seed is provided and not 0
            if opt_seed is not None and opt_seed != 0:
                random.seed(opt_seed)
            selected_indices = [random.randint(0, len(rows) - 1)]
        elif selection_mode == "multiple":
            if selected_row:
                try:
                    selected_indices = [int(idx) for idx in selected_row.split(",") if idx.strip().isdigit() and int(idx) < len(rows)]
                except Exception:
                    selected_indices = []
        else:  # single
            if not selected_row or not selected_row.isdigit() or int(selected_row) >= len(rows):
                selected_indices = [0]
            else:
                selected_indices = [int(selected_row)]

        # Format output for main output
        outputs = []
        for idx in selected_indices:
            row = rows[idx]
            out = ""
            for h, v in zip(headers, row):
                out += f"{h}:\n{v}\n\n"
            outputs.append(out.strip())
        output_str = "\n---\n".join(outputs)

        # Format output for ALL_SELECTED_ROWS (list)
        if len(selected_indices) <= 1:
            all_indices = list(range(len(rows)))
        else:
            all_indices = selected_indices
        all_outputs = []
        for idx in all_indices:
            row = rows[idx]
            out = ""
            for h, v in zip(headers, row):
                out += f"{h}:\n{v}\n\n"
            all_outputs.append(out.strip())

        return (output_str, xlsx_file, all_outputs)

    @classmethod
    def IS_CHANGED(cls, xlsx_file, selection_mode, sheet_name="", selected_row="", filter_text="", opt_seed=0):
        if selection_mode == "random":
            # For random mode, include seed in the hash only if opt_seed is provided and not 0
            if opt_seed is not None and opt_seed != 0:
                return str(opt_seed) + str(xlsx_file) + str(selection_mode) + str(filter_text) + str(sheet_name)
            else:
                return float('nan')  # Fall back to normal random behavior
        return selected_row + str(xlsx_file) + str(selection_mode) + str(sheet_name)

    @classmethod
    def VALIDATE_INPUTS(cls, xlsx_file, selection_mode="single", sheet_name="", selected_row="", filter_text="", opt_seed=0):
        global xlsx_path
        xlsx_file = os.path.join(xlsx_path, xlsx_file)
        xlsx_file = os.path.abspath(xlsx_file)
        if not os.path.isfile(xlsx_file):
            return "XLSX file does not exist"
        if not OPENPYXL_AVAILABLE:
            return "openpyxl is not installed. Run: pip install openpyxl"
        return True


def get_directory_structure(path):
    structure = {"name": os.path.basename(path), "children": [], "path": path, "expanded": False}
    try:
        with os.scandir(path) as entries:
            for entry in entries:
                if entry.is_dir():
                    structure["children"].append(get_directory_structure(entry.path))
    except PermissionError:
        pass
    return structure


def _cell_to_str(value):
    if value is None:
        return ""
    return str(value)


def get_sheet_names(file_path):
    if not OPENPYXL_AVAILABLE:
        return []
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            return wb.sheetnames
        finally:
            wb.close()
    except Exception as e:
        print(f"Error reading sheet names from {file_path}: {e}")
        return []


def get_rows_from_xlsx(file_path, sheet_name="", filter_text=""):
    if not OPENPYXL_AVAILABLE:
        return {"headers": [], "rows": [], "sheets": []}
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheets = wb.sheetnames
            if sheet_name and sheet_name in sheets:
                ws = wb[sheet_name]
            else:
                ws = wb[sheets[0]] if sheets else wb.active

            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append([_cell_to_str(cell) for cell in row])

            if not all_rows or len(all_rows) < 2:
                return {"headers": [], "rows": [], "sheets": sheets}

            headers = all_rows[0]
            rows = all_rows[1:]

            # Filter out empty rows (rows where all cells are empty or whitespace)
            rows = [row for row in rows if any(cell.strip() for cell in row)]

            if filter_text:
                rows = [row for row in rows if any(filter_text.lower() in str(cell).lower() for cell in row)]

            return {"headers": headers, "rows": rows, "sheets": sheets}
        finally:
            wb.close()
    except Exception as e:
        print(f"Error reading XLSX file {file_path}: {e}")
        return {"headers": [], "rows": [], "sheets": []}


@PromptServer.instance.routes.post("/ez_xlsx_browser/get_directory_structure")
async def api_get_directory_structure_xlsx(request):
    try:
        data = await request.json()
        path = data.get("path", "./")
        filter_text = data.get("filter", "")
        sheet_name = data.get("sheet", "")

        if not os.path.isabs(path):
            path = os.path.abspath(path)

        if not os.path.exists(path):
            return web.json_response({"error": "Path does not exist"}, status=400)

        # If path is a file, get its directory
        if os.path.isfile(path):
            directory = os.path.dirname(path)
            structure = get_directory_structure(directory)
            xlsx_data = get_rows_from_xlsx(path, sheet_name=sheet_name, filter_text=filter_text)
        else:
            structure = get_directory_structure(path)
            xlsx_data = {"headers": [], "rows": [], "sheets": []}

        response_data = {
            "structure": structure,
            "headers": xlsx_data["headers"],
            "rows": xlsx_data["rows"],
            "sheets": xlsx_data.get("sheets", []),
        }
        return web.json_response(response_data)
    except Exception as e:
        return web.json_response({"error": str(e)}, status=500)


@PromptServer.instance.routes.post("/ez_xlsx_browser/get_file_info")
async def get_file_info_xlsx(request):
    try:
        data = await request.json()
        rel_path = data.get("relative_path", "")
        full_path = os.path.normpath(os.path.join(xlsx_path, rel_path))
        if not full_path.startswith(xlsx_path):
            return web.json_response({"error": "Invalid path"}, status=400)
        if not os.path.exists(full_path):
            return web.json_response({"error": "File not found"}, status=404)
        return web.json_response({
            "full_path": full_path,
            "sheets": get_sheet_names(full_path),
        })
    except Exception as e:
        return web.json_response({"error": str(e)}, status=500)
